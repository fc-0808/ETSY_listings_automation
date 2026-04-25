"""Read the allowed_values sheet from Y2KASEshop.xlsx and map column letters to header + values."""
import openpyxl
from openpyxl.utils import get_column_letter

path = r"C:\Users\w088s\OneDrive\Documents\E-Commerce\Etsy\ETSY_listings_automation\Y2KASEshop.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)

template_ws = wb["Template"]
headers = {get_column_letter(i): cell.value for i, cell in enumerate(template_ws[1], 1) if cell.value}

allowed_ws = wb["allowed_values"]

# Build column-letter → list of allowed values
col_values: dict[str, list] = {}
for col_idx in range(1, allowed_ws.max_column + 1):
    col_letter = get_column_letter(col_idx)
    vals = []
    for row_idx in range(1, allowed_ws.max_row + 1):
        cell = allowed_ws.cell(row=row_idx, column=col_idx)
        v = cell.value
        if v is not None:
            vals.append(str(v).encode("ascii", "replace").decode("ascii"))
    if vals:
        col_values[col_letter] = vals

# Print with header name
print("Columns in allowed_values sheet with values:")
for col_letter, vals in sorted(col_values.items()):
    header = headers.get(col_letter, f"[unknown col {col_letter}]")
    print(f"\n{col_letter} → {header!r}")
    for v in vals[:30]:
        print(f"    {v!r}")
    if len(vals) > 30:
        print(f"    ... ({len(vals)} total)")
