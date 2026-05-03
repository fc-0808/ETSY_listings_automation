"""Build Shop Uploader–compatible XLSX files from processed product packages."""
from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from src.config import Config
from src.models import ProductPackage

log = logging.getLogger(__name__)

# Fallback if template file is missing (same as standard Shop Uploader export).
DEFAULT_COLUMNS: list[str] = [
    "listing_id", "parent_sku", "sku",
    "title", "description",
    "price", "quantity",
    "category",
    "_primary_color", "_secondary_color", "_occasion", "_holiday",
    "_deprecated_diameter", "_deprecated_dimensions", "_deprecated_fabric",
    "_deprecated_finish", "_deprecated_flavor", "_deprecated_height",
    "_deprecated_length", "_deprecated_material", "_deprecated_pattern",
    "_deprecated_scent", "_deprecated_size", "_deprecated_style",
    "_deprecated_weight", "_deprecated_width", "_deprecated_device",
    "option1_name", "option1_value",
    "option2_name", "option2_value",
    "image_1", "image_2", "image_3", "image_4", "image_5",
    "image_6", "image_7", "image_8", "image_9", "image_10",
    "shipping_profile_id", "readiness_state_id", "return_policy_id",
    "length", "width", "height", "dimensions_unit",
    "weight", "weight_unit",
    "type",
    "who_made", "is_made_to_order", "year_made",
    "is_vintage", "is_supply", "is_taxable", "auto_renew",
    "is_customizable", "is_personalizable", "personalization_is_required",
    "personalization_instructions", "personalization_char_count_max",
    "style_1", "style_2",
    "tag_1", "tag_2", "tag_3", "tag_4", "tag_5", "tag_6", "tag_7",
    "tag_8", "tag_9", "tag_10", "tag_11", "tag_12", "tag_13",
    "action", "listing_state", "overwrite_images",
]

_VIDEO_COL = "video_1"  # Shop Uploader column for listing video URL

_cached_columns: tuple[Path, list[str]] | None = None


def load_output_columns(cfg: Config) -> list[str]:
    """
    Read header row from cfg.template_path so each Etsy shop's generated template
    drives column order and any extra category columns.
    """
    global _cached_columns
    path = cfg.template_path.resolve()
    if _cached_columns and _cached_columns[0] == path:
        return list(_cached_columns[1])

    if not path.exists():
        log.warning("Shop Uploader template not found: %s — using DEFAULT_COLUMNS", path)
        cols = list(DEFAULT_COLUMNS)
        _cached_columns = (path, cols)
        return cols

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not row1:
            log.warning("Empty template: %s — using DEFAULT_COLUMNS", path)
            cols = list(DEFAULT_COLUMNS)
        else:
            cols = [str(c).strip() if c is not None else "" for c in row1]
            while cols and cols[-1] == "":
                cols.pop()
            if not cols:
                cols = list(DEFAULT_COLUMNS)
    finally:
        wb.close()

    unknown = [c for c in cols if c and c not in _compiler_column_keys()]
    if unknown:
        log.info(
            "Template has %d column(s) not filled by the compiler (left blank): %s",
            len(unknown),
            ", ".join(unknown[:12]) + ("…" if len(unknown) > 12 else ""),
        )

    _cached_columns = (path, cols)
    log.info("Loaded %d columns from template: %s", len(cols), path.name)
    return list(cols)


def clear_column_cache() -> None:
    """For tests or switching template mid-process."""
    global _cached_columns
    _cached_columns = None


def _safe_label(name: str) -> str:
    import re
    return re.sub(r"[^A-Za-z0-9_-]", "_", name.strip())[:40]


def _compiler_column_keys() -> frozenset[str]:
    return frozenset(DEFAULT_COLUMNS)


_HEADER_FILL = PatternFill("solid", fgColor="1a1a2e")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)


def build_xlsx(
    packages: list[ProductPackage],
    cfg: Config,
    batch_label: str = "",
) -> tuple[Path, int, list[str]]:
    """
    Write one XLSX file for Shop Uploader.

    Returns (output_path, rows_written, warnings).
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    shop = _safe_label(cfg.run_label or cfg.template_path.stem)
    state = "DRAFT" if cfg.listing_state == "draft" else "LIVE"
    batch_part = f"_{batch_label}" if batch_label else ""
    filename = f"UPLOAD_TO_SHOPUPLOADER__{shop}__{state}__{timestamp}{batch_part}.xlsx"
    output_path = cfg.output_dir / filename

    cols = load_output_columns(cfg)
    # Inject columns that may not exist in older templates
    for extra_col, before_col in [
        (_VIDEO_COL, "action"),
        ("linked_image_url", "action"),
        ("linked_image_position", "action"),
        ("linked_image_for_option", "action"),
        ("linked_image_alt_text", "action"),   # SEO: alt text for linked image
        ("production_partner_1", "action"),
        ("shop_section_id", "action"),
        ("featured_rank", "action"),
        # Explicitly set FALSE so "Processing profiles vary" stays OFF
        ("option1_changes_readiness_state", "option1_name"),
        ("option2_changes_readiness_state", "option2_name"),
        # Per Shop Uploader variations_advanced docs:
        # option1 = Phone Model → FALSE for all (price/qty/sku do NOT change per phone model)
        # option2 = Styles      → TRUE  for all (price/qty/sku DO change per style)
        ("option1_changes_price",    "option1_name"),
        ("option1_changes_quantity", "option1_name"),
        ("option1_changes_sku",      "option1_name"),
        ("option2_changes_price",    "option2_name"),
        ("option2_changes_quantity", "option2_name"),
        ("option2_changes_sku",      "option2_name"),
        # Controls whether a variation is purchaseable / visible to buyers
        ("variation_is_enabled",     "option1_name"),
    ]:
        if extra_col not in cols:
            insert_at = cols.index(before_col) if before_col in cols else len(cols)
            cols.insert(insert_at, extra_col)
            log.info("Injected '%s' column before '%s'", extra_col, before_col)
    columns = cols

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"

    _write_header(ws, columns)

    rows_written = 0
    warnings: list[str] = []

    for pkg in packages:
        if not pkg.is_ready:
            warnings.append(
                f"[{pkg.meta.parent_sku}] Package not ready "
                f"(images={bool(pkg.image_urls)}, copy={pkg.generated_copy is not None}) — skipped"
            )
            continue

        product_rows = _build_rows(pkg, cfg, columns)
        for row in product_rows:
            ws.append(row)
        rows_written += len(product_rows)

    _auto_fit_columns(ws, columns)
    wb.save(str(output_path))
    log.info("XLSX written: %s (%d rows)", output_path.name, rows_written)

    return output_path, rows_written, warnings


def build_batched_xlsx_files(
    packages: list[ProductPackage],
    cfg: Config,
) -> tuple[list[Path], list[str]]:
    """
    Split packages into cfg.batch_size chunks and write one file per batch.
    Returns (list_of_paths, all_warnings).
    """
    all_paths: list[Path] = []
    all_warnings: list[str] = []

    ready = [p for p in packages if p.is_ready]
    total_batches = max(1, -(-len(ready) // cfg.batch_size))

    for batch_num in range(total_batches):
        start = batch_num * cfg.batch_size
        end = start + cfg.batch_size
        batch = ready[start:end]
        if not batch:
            continue
        label = f"batch{batch_num + 1:03d}_of_{total_batches:03d}"
        path, rows_written, warnings = build_xlsx(batch, cfg, batch_label=label)
        all_paths.append(path)
        all_warnings.extend(warnings)
        log.info("Batch %d/%d: %d rows → %s", batch_num + 1, total_batches, rows_written, path.name)

    return all_paths, all_warnings


def _order_listing_images(
    image_urls: list[str],
) -> list[str]:
    """Return image URLs in the order the seller set by renaming files."""
    return [u for u in image_urls if u][:10]


def _derive_mapping_from_analysis(
    image_analysis: list[dict],
    image_urls: list[str],
) -> dict[str, list[int]]:
    """
    Fallback: build a style → [1-based image indices] mapping from the per-image
    boolean facts that the AI returned in image_analysis.

    This runs when style_image_mapping is all-empty (AI classified images but
    didn't fill the mapping, e.g. due to model limitations).
    Index 1 is the thumbnail and is intentionally excluded from linked images.
    """
    mapping: dict[str, list[int]] = {
        "Case+Grip+Charm": [], "Case+Grip": [], "Case+Charm": [],
        "Case Only": [], "Grip Only": [], "Charm Only": [],
    }
    n = len(image_urls)
    for img in sorted(image_analysis, key=lambda x: x.get("index", 0)):
        idx = img.get("index", 0)
        if idx < 2 or idx > n:   # skip index 1 (thumbnail) and out-of-range
            continue
        has_grip  = bool(img.get("has_grip",  False))
        has_charm = bool(img.get("has_charm", False))
        has_case  = bool(img.get("has_case",  True))
        edge      = bool(img.get("is_edge_or_profile", False))
        if has_case:
            if has_grip and has_charm:
                mapping["Case+Grip+Charm"].append(idx)
            elif has_grip:
                mapping["Case+Grip"].append(idx)
            elif has_charm:
                mapping["Case+Charm"].append(idx)
            elif not edge:
                mapping["Case Only"].append(idx)
        # standalone grip or charm — omit from linked images
    # Keep only styles with at least one image
    return {k: v for k, v in mapping.items() if v}


def _build_image_alt_texts(
    image_analysis: list[dict],
    parent_sku: str,
) -> dict[str, str]:
    """
    Build image_alt_text_1..10 from Phase 1 per-image descriptions.
    Per Shop Uploader docs: image_alt_text_N corresponds to image_N and
    improves Etsy/Google SEO. We use the Phase 1 description as the base
    and truncate to a sensible length (max 250 chars per Etsy guideline).
    """
    result: dict[str, str] = {}
    if not image_analysis:
        return result
    # Build index → description lookup
    desc_by_idx = {
        img.get("index", 0): img.get("description", "")
        for img in image_analysis
        if img.get("description")
    }
    for pos in range(1, 11):   # image_1 through image_10
        desc = desc_by_idx.get(pos, "")
        if desc:
            result[f"image_alt_text_{pos}"] = desc[:250].strip()
    return result


def _get_alt_text_for_linked(
    linked_url: str,
    image_urls: list[str],
    image_analysis: list[dict],
) -> str:
    """
    Return Phase 1 description for the image that was linked, trimmed for alt text.
    Matches by URL position to get the correct 1-based index.
    """
    if not linked_url or not image_analysis:
        return ""
    try:
        pos = image_urls.index(linked_url) + 1  # 1-based
    except ValueError:
        return ""
    for img in image_analysis:
        if img.get("index") == pos:
            return str(img.get("description", ""))[:250].strip()
    return ""


def _write_header(ws, columns: list[str]) -> None:
    ws.append(columns)
    for col_idx, _ in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _build_rows(pkg: ProductPackage, cfg: Config, columns: list[str]) -> list[list]:
    """
    Return a list of spreadsheet rows for one product.
    - No variations → 1 row.
    - With variations (Style × Phone Model) → len(styles) × len(models) rows.
      All rows share the same parent_sku; the first row carries the full listing data.
    """
    meta = pkg.meta
    copy = pkg.generated_copy
    tags = (copy.tags + [""] * 13)[:13]

    # ── Image ordering — seller-defined by filename order ─────────────────────
    raw_urls = list(pkg.image_urls)
    ordered_urls = _order_listing_images(raw_urls)
    urls = (ordered_urls + [""] * 10)[:10]
    # ── Base listing fields (shared by every row in this product) ─────────────
    listing_fields: dict[str, object] = {
        "listing_id": "",
        "parent_sku": meta.parent_sku,
        "title": copy.title,
        "description": copy.description,
        "category": meta.category,
        "video_1": pkg.video_url,  # empty string if no video
        # Colors: AI-detected from images takes priority over meta.json defaults
        "_primary_color": copy.primary_color or meta.category_properties.primary_color,
        "_secondary_color": copy.secondary_color or meta.category_properties.secondary_color,
        "_occasion": meta.category_properties.occasion,
        "_holiday": meta.category_properties.holiday,
        # Multi-material attribute (new template column replacing legacy _material)
        "_material_multi": meta.category_properties.material,
        # Phone Case category feature attributes — read from meta.json category_properties
        "_built-in_grip":          meta.category_properties.built_in_grip,
        "_built-in_stand":         meta.category_properties.built_in_stand,
        "_card_slot":              meta.category_properties.card_slot,
        "_electronics_case_theme": meta.category_properties.electronics_case_theme,
        "_glitter":                meta.category_properties.glitter,
        "_liquid":                 meta.category_properties.liquid,
        "_pattern":                meta.category_properties.pattern,
        "_deprecated_diameter": "", "_deprecated_dimensions": "",
        "_deprecated_fabric": "", "_deprecated_finish": "",
        "_deprecated_flavor": "", "_deprecated_height": "",
        "_deprecated_length": "", "_deprecated_material": "",
        "_deprecated_pattern": "", "_deprecated_scent": "",
        "_deprecated_size": "", "_deprecated_style": "",
        "_deprecated_weight": "", "_deprecated_width": "",
        "_deprecated_device": "",
        "image_1": urls[0], "image_2": urls[1], "image_3": urls[2],
        "image_4": urls[3], "image_5": urls[4], "image_6": urls[5],
        "image_7": urls[6], "image_8": urls[7], "image_9": urls[8],
        "image_10": urls[9],
        # Alt text per numbered image — sourced from Phase 1 per-image descriptions.
        # Improves Etsy SEO (image alt text is indexed by Etsy and Google Shopping).
        # Per Shop Uploader docs: image_alt_text_N corresponds to image_N.
        **_build_image_alt_texts(copy.image_analysis, meta.parent_sku),
        "shipping_profile_id": meta.shipping_profile_id,
        "readiness_state_id": meta.readiness_state_id,
        "return_policy_id": meta.return_policy_id,
        "length": meta.length, "width": meta.width, "height": meta.height,
        "dimensions_unit": meta.dimensions_unit,
        "weight": meta.weight, "weight_unit": meta.weight_unit,
        "type": meta.type,
        "who_made": meta.who_made,
        "is_made_to_order": _bool(meta.is_made_to_order),
        "year_made": meta.year_made,
        "is_vintage": _bool(meta.is_vintage),
        "is_supply": _bool(meta.is_supply),
        "is_taxable": _bool(meta.is_taxable),
        "auto_renew": _bool(meta.auto_renew),
        "is_customizable": _bool(meta.is_customizable),
        "is_personalizable": _bool(meta.is_personalizable),
        "personalization_is_required": _bool(meta.personalization_is_required),
        "personalization_instructions": meta.personalization_instructions,
        "personalization_char_count_max": meta.personalization_char_count_max,
        "style_1": meta.style_1, "style_2": meta.style_2,
        "tag_1": tags[0], "tag_2": tags[1], "tag_3": tags[2],
        "tag_4": tags[3], "tag_5": tags[4], "tag_6": tags[5],
        "tag_7": tags[6], "tag_8": tags[7], "tag_9": tags[8],
        "tag_10": tags[9], "tag_11": tags[10], "tag_12": tags[11],
        "tag_13": tags[12],
        "action": "create",
        "listing_state": cfg.listing_state,
        "overwrite_images": "TRUE",
        # linked image columns — populated per variation row below
        "linked_image_url": "",
        "linked_image_position": "",
        "linked_image_for_option": "",
        "linked_image_alt_text": "",
        # Production partner — same for every listing in this shop
        "production_partner_1": meta.production_partner_1,
        # Shop section — "iPhone Cases" section
        "shop_section_id": meta.shop_section_id,
        # Featured rank — any positive integer enables "Feature this listing"
        "featured_rank": meta.featured_rank,
        # Explicitly FALSE so "Processing profiles vary" stays OFF in Etsy
        "option1_changes_readiness_state": "FALSE",
        "option2_changes_readiness_state": "FALSE",
        # Variations advanced — Phone Model (option1) drives NOTHING
        "option1_changes_price":    "FALSE",
        "option1_changes_quantity": "FALSE",
        "option1_changes_sku":      "FALSE",
        # Styles (option2) drives EVERYTHING
        "option2_changes_price":    "TRUE",
        "option2_changes_quantity": "TRUE",
        "option2_changes_sku":      "TRUE",
        # Enabled by default; disabled per-row for styles the product doesn't offer
        "variation_is_enabled":     "TRUE",
    }

    # ── Determine which styles this product actually offers ────────────────────
    # Primary signal: style_image_mapping — if the AI mapped at least one image to
    # a grip-containing style then the product sells grip bundles.
    # A non-empty dict with ALL-EMPTY lists means AI returned no classifications,
    # so fall through to image_analysis regardless.
    sim = copy.style_image_mapping or {}
    _sim_has_data = any(bool(v) for v in sim.values())  # True only if ANY style has images
    ia = copy.image_analysis or []

    if _sim_has_data:
        product_has_grip  = bool(sim.get("Case+Grip", []) or sim.get("Case+Grip+Charm", []))
        product_has_charm = bool(sim.get("Case+Charm", []) or sim.get("Case+Grip+Charm", []))
        # Belt-and-suspenders: if the style map shows no grip/charm but the per-image
        # analysis explicitly flags them (has_grip/has_charm=True), trust the image
        # analysis over the mapping — the mapping may have missed an edge case.
        ia_has_grip  = any(img.get("has_grip",  False) for img in ia)
        ia_has_charm = any(img.get("has_charm", False) for img in ia)
        if not product_has_grip and ia_has_grip:
            log.warning(
                "[%s] style_map shows no grip images but image_analysis disagrees "
                "— overriding product_has_grip to True",
                meta.parent_sku,
            )
            product_has_grip = True
        if not product_has_charm and ia_has_charm:
            log.warning(
                "[%s] style_map shows no charm images but image_analysis disagrees "
                "— overriding product_has_charm to True",
                meta.parent_sku,
            )
            product_has_charm = True
    else:
        # AI didn't classify any images — derive from per-image boolean facts
        product_has_grip  = any(img.get("has_grip",  False) for img in ia)
        product_has_charm = any(img.get("has_charm", False) for img in ia)

    # Styles that require a grip to exist; hidden when product has no grip
    _GRIP_REQUIRED_STYLES  = {"Case+Grip+Charm", "Case+Grip", "Grip Only"}
    # Styles that require a charm to exist; hidden when product has no charm
    _CHARM_REQUIRED_STYLES = {"Case+Grip+Charm", "Case+Charm", "Charm Only"}

    def _variation_enabled(style_value: str) -> str:
        if not product_has_grip  and style_value in _GRIP_REQUIRED_STYLES:
            return "FALSE"
        if not product_has_charm and style_value in _CHARM_REQUIRED_STYLES:
            return "FALSE"
        return "TRUE"

    # Diagnostic log — shows the final enabled/disabled state for every style so
    # any mismatch with Etsy is immediately visible in the run log.
    _ALL_STYLE_VALUES = [
        "Case+Grip+Charm", "Case+Grip", "Case+Charm",
        "Case Only", "Grip Only", "Charm Only",
    ]
    log.info(
        "[%s] Variation enablement (grip=%s charm=%s): %s",
        meta.parent_sku,
        product_has_grip,
        product_has_charm,
        "  ".join(
            f"{s}={'ON' if _variation_enabled(s) == 'TRUE' else 'OFF'}"
            for s in _ALL_STYLE_VALUES
        ),
    )

    v = meta.variations
    if v is None:
        # ── No variations: single row, use top-level price/qty/sku ───────────
        row = dict(listing_fields)
        row["sku"] = meta.sku
        row["price"] = meta.price
        row["quantity"] = meta.quantity
        row["option1_name"] = ""
        row["option1_value"] = ""
        row["option2_name"] = ""
        row["option2_value"] = ""
        return [[row.get(col, "") for col in columns]]

    # ── 2D variations: one row per (model × style) combination ──────────────
    # option1 = Phone Model (12), option2 = Styles (6) → 72 rows per product
    # Prices/Quantities/SKUs driven by Styles only.
    # Linked images: AI determined which image best represents each Style.
    rows: list[list] = []
    first_row = True

    # Build style → image index list mapping.
    # Priority 1: AI-provided style_image_mapping (only keep styles with ≥1 image).
    # Priority 2: Derive from per-image image_analysis boolean facts.
    # "Case Only (edge)" images are merged as low-priority fallback for "Case Only".
    sim = copy.style_image_mapping or {}
    ai_mapping: dict[str, list[int]] = {k: v for k, v in sim.items() if v}

    # Merge "Case Only (edge)" into "Case Only" when no nice-angle case images exist
    edge_indices: list[int] = sim.get("Case Only (edge)") or []  # type: ignore[assignment]
    if isinstance(edge_indices, int):
        edge_indices = [edge_indices]
    if edge_indices and not ai_mapping.get("Case Only"):
        ai_mapping["Case Only"] = edge_indices

    # Fallback: derive from image_analysis booleans when AI mapping is all-empty
    if not ai_mapping and copy.image_analysis:
        ai_mapping = _derive_mapping_from_analysis(copy.image_analysis, pkg.image_urls)
        if ai_mapping:
            log.info(
                "[%s] style_image_mapping was empty — derived from image_analysis: %s",
                meta.parent_sku,
                {k: v for k, v in ai_mapping.items()},
            )

    # Build style_name → (url, position) lookup.
    # RULE: never use index 1 (the thumbnail) as a linked image.
    # The linked_image_position MUST equal the image's own 1-based index in image_urls
    # so that Shop Uploader's deduplication collapses it at exactly the right slot,
    # preserving the seller's file order throughout all positions.
    # Charm Only is ALWAYS excluded from linked images.
    def get_linked_url_and_pos(style_name: str) -> tuple[str, str]:
        if style_name == "Charm Only":
            return "", ""
        indices = ai_mapping.get(style_name, [])
        if isinstance(indices, int):
            indices = [indices]
        for idx in indices:
            if idx == 1:
                continue  # never link to index 1 — it is the thumbnail
            zero = idx - 1
            if 0 <= zero < len(pkg.image_urls):
                return pkg.image_urls[zero], str(idx)  # position = image's own 1-based index
        return "", ""

    for model in v.models:
        for style in v.styles:
            safe_model = model.replace(" ", "").replace("Pro", "P").replace("Max", "M").replace("/", "")
            sku = f"{meta.parent_sku}-{style.sku_code}-{safe_model}"[:32]

            row = dict(listing_fields)
            # SKU: blank to match live listing (Etsy auto-assigns variation IDs).
            # This is what makes "SKUs vary for Styles only" — when SKU is blank,
            # Etsy ties SKU dimension to the same dimension as price (Styles).
            row["sku"] = ""
            row["price"] = style.price
            row["quantity"] = style.quantity

            # option1 = Phone Model, option2 = Styles (matches live Y2KASEshop)
            row["option1_name"] = v.option1_name   # "Phone Model"
            row["option1_value"] = model
            row["option2_name"] = v.option2_name   # "Styles"
            row["option2_value"] = style.value

            # Linked image position = the image's own sequential index.
            # Per Shop Uploader docs, positions are relative — setting position N
            # to match the same URL as image_N causes Shop Uploader to deduplicate
            # rather than insert the image twice, preserving the seller's file order.
            linked_url, link_pos = get_linked_url_and_pos(style.value)
            row["linked_image_url"]      = linked_url
            row["linked_image_position"] = link_pos

            # ── CRITICAL: linked_image_for_option must ALWAYS be set to enable ──
            # "Link photos to this variation" in Shop Uploader / Etsy.
            # Per official docs: this column holds the OPTION NAME ("Styles"),
            # not the option value. Even rows without a linked image must carry
            # this name so the toggle stays ON for the whole listing.
            row["linked_image_for_option"] = v.option2_name  # always "Styles"

            # Alt text for the linked image — use Phase 1 description of the
            # best image for this style. Good for Etsy/Google SEO.
            row["linked_image_alt_text"] = (
                _get_alt_text_for_linked(linked_url, pkg.image_urls, copy.image_analysis)
                if linked_url else ""
            )

            # Disable variation for styles the product doesn't offer
            row["variation_is_enabled"] = _variation_enabled(style.value)

            if not first_row:
                for blank_col in (
                    "title", "description", "video_1",
                    "image_1", "image_2", "image_3", "image_4", "image_5",
                    "image_6", "image_7", "image_8", "image_9", "image_10",
                    # alt texts only on first row (same as numbered images)
                    "image_alt_text_1", "image_alt_text_2", "image_alt_text_3",
                    "image_alt_text_4", "image_alt_text_5", "image_alt_text_6",
                    "image_alt_text_7", "image_alt_text_8", "image_alt_text_9",
                    "image_alt_text_10",
                    "tag_1", "tag_2", "tag_3", "tag_4", "tag_5", "tag_6",
                    "tag_7", "tag_8", "tag_9", "tag_10", "tag_11", "tag_12", "tag_13",
                ):
                    row[blank_col] = ""
            first_row = False

            rows.append([row.get(col, "") for col in columns])

    return rows


def _build_row(pkg: ProductPackage, cfg: Config, columns: list[str]) -> list:
    """Kept for backward-compat; delegates to _build_rows."""
    return _build_rows(pkg, cfg, columns)[0]


def _bool(value: bool) -> str:
    return "TRUE" if value else "FALSE"


def _bool_to_excel(value: bool) -> str:
    return "TRUE" if value else "FALSE"


def _auto_fit_columns(ws, columns: list[str]) -> None:
    for col_idx in range(1, len(columns) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    cell_len = len(str(cell.value or ""))
                    if cell_len > max_len:
                        max_len = cell_len
                except Exception:
                    pass
        adjusted_width = min(max(max_len + 2, 12), 60)
        ws.column_dimensions[col_letter].width = adjusted_width
